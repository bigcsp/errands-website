#!/usr/bin/env python3
"""
Errands — Orders Excel Agent
─────────────────────────────
Usage:
  python agent.py                        # uses errands_orders_*.json in current dir
  python agent.py orders.json            # specific file
  python agent.py orders.json out.xlsx   # custom output filename

What it does:
  Reads the JSON exported from admin.html and generates a
  beautifully formatted errands_orders_<date>.xlsx file.
"""

import json, sys, os, glob
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠  openpyxl not found. Installing…")
    os.system("pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─── COLOURS ─────────────────────────────────────────────────────────────────
C_ORANGE      = "FF6B2B"
C_DARK        = "0A0A0A"
C_HEADER_BG   = "1A1A1A"
C_ROW_ALT     = "111111"
C_ROW_NORMAL  = "0D0D0D"
C_WHITE       = "F0F0F0"
C_MUTED       = "888888"
C_GREEN       = "22C55E"
C_YELLOW      = "F59E0B"
C_BLUE        = "3B82F6"
C_RED         = "EF4444"
C_PURPLE      = "A855F7"

STATUS_COLORS = {
    "Pending":    C_YELLOW,
    "In Transit": C_BLUE,
    "Nearby":     C_PURPLE,
    "Delivered":  C_GREEN,
    "Cancelled":  C_RED,
}


def thin_border(color="333333"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, cell_ref, value=None, bold=False, font_size=11,
               font_color=C_WHITE, bg_color=None, align="left",
               wrap=False, border=False):
    c = ws[cell_ref]
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", bold=bold, size=font_size, color=font_color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg_color:
        c.fill = PatternFill("solid", fgColor=bg_color)
    if border:
        c.border = thin_border()
    return c


# ─── LOAD JSON ────────────────────────────────────────────────────────────────
def load_orders(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ─── SHEET 1: ORDERS ─────────────────────────────────────────────────────────
COLUMNS = [
    ("Tracking ID",       "id",              16),
    ("Status",            "status",          13),
    ("Order Date",        "createdAt",       20),
    ("Sender Name",       "senderName",      18),
    ("Sender Phone",      "senderPhone",     15),
    ("Sender Email",      "senderEmail",     22),
    ("Pickup Address",    "pickupAddress",   32),
    ("Pickup Date",       "pickupDate",      13),
    ("Pickup Time",       "pickupTime",      12),
    ("Drop-off Address",  "dropoffAddress",  32),
    ("Recipient Name",    "recipientName",   18),
    ("Recipient Phone",   "recipientPhone",  15),
    ("Package Type",      "packageType",     18),
    ("Weight (kg)",       "packageWeight",   12),
    ("Price (NGN)",       "price",           13),
    ("Notes",             "notes",           28),
    ("Rider",             "rider",           14),
]


def build_orders_sheet(wb, orders):
    ws = wb.active
    ws.title = "Orders"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # ── Title row ──
    ws.merge_cells("A1:Q1")
    c = ws["A1"]
    c.value = f"ERRANDS — DELIVERY ORDERS   ·   Generated {datetime.now().strftime('%d %b %Y, %H:%M')}"
    c.font      = Font(name="Arial", bold=True, size=13, color=C_ORANGE)
    c.fill      = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header row ──
    ws.row_dimensions[2].height = 22
    for col_i, (label, _, width) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_i)
        ws.column_dimensions[col_letter].width = width
        c = ws.cell(row=2, column=col_i, value=label.upper())
        c.font      = Font(name="Arial", bold=True, size=9, color=C_ORANGE)
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border("222222")

    # ── Data rows ──
    for row_i, order in enumerate(orders, start=3):
        ws.row_dimensions[row_i].height = 18
        bg = C_ROW_ALT if row_i % 2 == 0 else C_ROW_NORMAL

        for col_i, (_, field, _) in enumerate(COLUMNS, start=1):
            raw = order.get(field, "")

            # Format createdAt
            if field == "createdAt" and raw:
                try:
                    raw = datetime.fromisoformat(raw.replace("Z","")).strftime("%d %b %Y, %H:%M")
                except Exception:
                    pass

            # Format price
            if field == "price" and raw:
                raw = f"₦{int(raw):,}"

            c = ws.cell(row=row_i, column=col_i, value=str(raw) if raw else "")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = thin_border("1A1A1A")

            font_color = C_WHITE

            # Tracking ID highlight
            if field == "id":
                font_color = C_ORANGE
                c.font = Font(name="Arial", bold=True, size=10, color=C_ORANGE)
                continue

            # Status colour
            if field == "status":
                sc = STATUS_COLORS.get(str(raw), C_MUTED)
                c.font = Font(name="Arial", bold=True, size=9, color=sc)
                continue

            # Price colour
            if field == "price":
                c.font = Font(name="Arial", bold=True, size=10, color=C_GREEN)
                continue

            c.font = Font(name="Arial", size=10, color=font_color)

    # ── Auto-filter ──
    ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"


# ─── SHEET 2: SUMMARY ────────────────────────────────────────────────────────
def build_summary_sheet(wb, orders):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    total     = len(orders)
    pending   = sum(1 for o in orders if o.get("status") == "Pending")
    in_transit= sum(1 for o in orders if o.get("status") == "In Transit")
    nearby    = sum(1 for o in orders if o.get("status") == "Nearby")
    delivered = sum(1 for o in orders if o.get("status") == "Delivered")
    cancelled = sum(1 for o in orders if o.get("status") == "Cancelled")
    revenue   = sum(o.get("price", 0) for o in orders if o.get("status") == "Delivered")

    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = "ERRANDS — SUMMARY DASHBOARD"
    c.font      = Font(name="Arial", bold=True, size=14, color=C_ORANGE)
    c.fill      = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws["A2"].value = f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws["A2"].font  = Font(name="Arial", size=9, color=C_MUTED)
    ws.row_dimensions[2].height = 16

    rows_data = [
        ("",            "",          C_DARK,       C_WHITE,  False),
        ("METRIC",      "VALUE",     C_HEADER_BG,  C_ORANGE, True),
        ("Total Orders",total,       C_ROW_NORMAL, C_WHITE,  False),
        ("Pending",     pending,     C_ROW_ALT,    C_YELLOW, False),
        ("In Transit",  in_transit,  C_ROW_NORMAL, C_BLUE,   False),
        ("Nearby",      nearby,      C_ROW_ALT,    C_PURPLE, False),
        ("Delivered",   delivered,   C_ROW_NORMAL, C_GREEN,  False),
        ("Cancelled",   cancelled,   C_ROW_ALT,    C_RED,    False),
        ("",            "",          C_DARK,       C_WHITE,  False),
        ("Total Revenue (NGN)", f"₦{revenue:,}", C_ROW_NORMAL, C_GREEN, True),
        ("Delivery Rate", f"{round(delivered/total*100 if total else 0, 1)}%", C_ROW_ALT, C_WHITE, False),
    ]

    for r_i, (label, val, bg, fc, bold) in enumerate(rows_data, start=3):
        ws.row_dimensions[r_i].height = 20
        for col_i, v in enumerate([label, val], start=1):
            c = ws.cell(row=r_i, column=col_i, value=v)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(name="Arial", bold=bold, size=11, color=fc if col_i == 2 else C_WHITE)
            c.alignment = Alignment(horizontal="left" if col_i == 1 else "right", vertical="center")
            c.border    = thin_border("1A1A1A")


# ─── SHEET 3: PIVOT by STATUS ─────────────────────────────────────────────────
def build_pivot_sheet(wb, orders):
    ws = wb.create_sheet("By Status")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "ORDERS BY STATUS"
    c.font      = Font(name="Arial", bold=True, size=13, color=C_ORANGE)
    c.fill      = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Status", "Count", "Revenue (₦)"]
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_i, value=h.upper())
        c.font      = Font(name="Arial", bold=True, size=9, color=C_ORANGE)
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border("222222")
    ws.row_dimensions[2].height = 20

    for r_i, status in enumerate(["Pending","In Transit","Nearby","Delivered","Cancelled"], start=3):
        group   = [o for o in orders if o.get("status") == status]
        count   = len(group)
        revenue = sum(o.get("price", 0) for o in group)
        color   = STATUS_COLORS.get(status, C_MUTED)
        bg      = C_ROW_ALT if r_i % 2 == 0 else C_ROW_NORMAL

        ws.row_dimensions[r_i].height = 18
        for col_i, v in enumerate([status, count, f"₦{revenue:,}"], start=1):
            c = ws.cell(row=r_i, column=col_i, value=v)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(name="Arial", bold=(col_i == 1), size=10, color=color if col_i == 1 else C_WHITE)
            c.alignment = Alignment(horizontal="left" if col_i == 1 else "center", vertical="center")
            c.border    = thin_border("1A1A1A")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    # Find input file
    if len(sys.argv) >= 2:
        input_path = sys.argv[1]
    else:
        matches = glob.glob("errands_orders_*.json") + glob.glob("orders.json")
        if not matches:
            print("❌  No JSON file found. Export one from admin.html first.")
            print("    Then run:  python agent.py errands_orders_YYYY-MM-DD.json")
            sys.exit(1)
        input_path = sorted(matches)[-1]
        print(f"📂  Auto-detected: {input_path}")

    if not os.path.exists(input_path):
        print(f"❌  File not found: {input_path}")
        sys.exit(1)

    # Output filename
    date_str    = datetime.now().strftime("%Y-%m-%d")
    output_path = sys.argv[2] if len(sys.argv) >= 3 else f"errands_orders_{date_str}.xlsx"

    print(f"⏳  Loading orders from {input_path}…")
    orders = load_orders(input_path)
    print(f"✅  {len(orders)} orders loaded")

    print("📊  Building Excel workbook…")
    wb = Workbook()
    build_orders_sheet(wb, orders)
    build_summary_sheet(wb, orders)
    build_pivot_sheet(wb, orders)

    wb.save(output_path)
    print(f"\n✅  Done! Saved to: {output_path}")
    print(f"    Sheets: Orders  |  Summary  |  By Status")
    print(f"    Orders: {len(orders)}  |  Open in Excel / LibreOffice")


if __name__ == "__main__":
    main()
